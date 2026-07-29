"""Byte-identical regeneration, and a test that cannot pass by coincidence.

Issue #13. Freezing `workbook.properties.modified` is not sufficient: openpyxl
re-stamps it on save regardless, and each ZIP local header carries a wall-clock
mtime. Measured: every member decompresses identically while the files hash
differently.

The naive check - save twice, compare hashes - passes whenever both saves land
in the same wall-clock second, which is most of the time. It would green-light a
broken fix. So the load-bearing assertions here are on the archive metadata
itself, not on a hash comparison that luck can satisfy.
"""

import hashlib
import time
import zipfile
import zlib
from pathlib import Path

import pytest
from openpyxl import Workbook

from procurement_agent.services import output as output_module
from procurement_agent.services.output import (
    DETERMINISTIC_APPLICATION,
    DETERMINISTIC_COMPRESSLEVEL,
    DETERMINISTIC_TIMESTAMP,
    normalize_archive,
)


def _write(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["A1"] = "PV Modules"
    sheet["B2"] = 650
    workbook.save(path)
    return path


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_every_archive_member_carries_the_fixed_timestamp(tmp_path: Path) -> None:
    """The assertion luck cannot satisfy: no member may keep a clock-derived mtime."""
    book = normalize_archive(_write(tmp_path / "book.xlsx"))
    with zipfile.ZipFile(book) as archive:
        stamps = {info.date_time for info in archive.infolist()}
    assert stamps == {DETERMINISTIC_TIMESTAMP}


def test_unnormalized_saves_are_not_byte_identical_in_general(tmp_path: Path) -> None:
    """Why normalisation is needed at all: the raw archive carries the clock.

    Asserted on the metadata rather than on hashes, because two raw saves inside
    one wall-clock second *do* hash the same - exactly the coincidence that makes
    the naive test unsound.
    """
    raw = _write(tmp_path / "raw.xlsx")
    with zipfile.ZipFile(raw) as archive:
        stamps = {info.date_time for info in archive.infolist()}
    assert stamps != {DETERMINISTIC_TIMESTAMP}


def test_normalized_saves_are_byte_identical_across_a_second_boundary(tmp_path: Path) -> None:
    """Deliberately slow. A save-twice-compare test that does NOT cross a second
    boundary passes on a broken fix - which is how the first attempt at this
    shipped green while `docProps/core.xml` still carried the wall clock."""
    first = normalize_archive(_write(tmp_path / "first.xlsx"))
    time.sleep(1.1)
    second = normalize_archive(_write(tmp_path / "second.xlsx"))
    assert _digest(first) == _digest(second)


def test_core_xml_timestamps_are_frozen(tmp_path: Path) -> None:
    """The member the first attempt missed entirely."""
    book = normalize_archive(_write(tmp_path / "book.xlsx"))
    with zipfile.ZipFile(book) as archive:
        core = archive.read("docProps/core.xml").decode()
    assert "1980-01-01T12:00:00Z" in core
    assert core.count("1980-01-01T12:00:00Z") >= 1
    for year in ("2024", "2025", "2026", "2027"):
        assert year not in core


def test_content_types_is_first_and_the_rest_sorted(tmp_path: Path) -> None:
    """plan.md Decision 8c. openpyxl writes it last."""
    book = normalize_archive(_write(tmp_path / "book.xlsx"))
    with zipfile.ZipFile(book) as archive:
        names = archive.namelist()
    assert names[0] == "[Content_Types].xml"
    assert names[1:] == sorted(names[1:])


def test_content_types_wins_against_a_member_that_sorts_before_it(tmp_path: Path) -> None:
    """The assertion above is a tautology on openpyxl's own output: every member
    it writes starts with `_`, `d` or `x`, and `[` (0x5B) sorts before all of
    them, so a plain `sorted()` already puts `[Content_Types].xml` first.

    A member sorting before `[` is what distinguishes the two, and the archive is
    non-conformant the moment the writer emits one."""
    book = _write(tmp_path / "book.xlsx")
    with zipfile.ZipFile(book, "a") as archive:
        archive.writestr("Custom/props.xml", b"<custom/>")
    with zipfile.ZipFile(normalize_archive(book)) as archive:
        names = archive.namelist()
    assert names[0] == "[Content_Types].xml"
    assert "Custom/props.xml" in names


def test_archive_metadata_does_not_vary_by_platform(tmp_path: Path) -> None:
    """`create_system` and `external_attr` otherwise differ on Windows."""
    book = normalize_archive(_write(tmp_path / "book.xlsx"))
    with zipfile.ZipFile(book) as archive:
        infos = archive.infolist()
    assert {i.create_system for i in infos} == {3}
    assert {i.external_attr for i in infos} == {0o644 << 16}
    assert {i.compress_type for i in infos} == {zipfile.ZIP_DEFLATED}


def test_create_system_is_forced_rather_than_left_to_the_platform(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """`ZipInfo.__init__` already defaults `create_system` to 3 on POSIX, so both
    the assertion above and a forged-fixture version of this test pass with the
    assignment deleted — `normalize_archive` builds fresh `ZipInfo` objects, so
    the source archive's value never reaches the output.

    What can be observed here is that the constant is *applied*: move it and the
    output moves with it. The value itself is pinned separately, because a test
    that only follows the constant would accept any value.
    """
    assert output_module._CREATE_SYSTEM_UNIX == 3, "Decision 8c: always Unix"
    monkeypatch.setattr(output_module, "_CREATE_SYSTEM_UNIX", 0)
    book = normalize_archive(_write(tmp_path / "book.xlsx"))
    with zipfile.ZipFile(book) as archive:
        assert {i.create_system for i in archive.infolist()} == {0}


def test_normalization_is_idempotent(tmp_path: Path) -> None:
    book = _write(tmp_path / "book.xlsx")
    once = _digest(normalize_archive(book))
    assert _digest(normalize_archive(book)) == once


def test_no_staging_file_is_left_behind(tmp_path: Path) -> None:
    normalize_archive(_write(tmp_path / "book.xlsx"))
    assert [p.name for p in tmp_path.iterdir()] == ["book.xlsx"]


def test_library_version_is_not_embedded_in_the_output(tmp_path: Path) -> None:
    """`docProps/app.xml` otherwise carries `Openpyxl <version>` verbatim, so a
    patch bump changes the bytes with zero data change and invalidates AC-7."""
    book = normalize_archive(_write(tmp_path / "book.xlsx"))
    with zipfile.ZipFile(book) as archive:
        app = archive.read("docProps/app.xml").decode()
    assert "Openpyxl" not in app
    assert DETERMINISTIC_APPLICATION in app
    # `<Application>` and `<AppVersion>` are two elements and openpyxl writes the
    # library version into both. Fixing only the first satisfied the two
    # assertions above while `<AppVersion>3.1</AppVersion>` still tracked the
    # library, so a 3.1 to 3.2 bump would change every workbook's bytes.
    assert "<AppVersion>1.0</AppVersion>" in app
    assert "3.1" not in app


def test_normalized_workbook_still_opens_and_keeps_its_data(tmp_path: Path) -> None:
    """Determinism is worthless if it corrupts the file."""
    from openpyxl import load_workbook

    book = normalize_archive(_write(tmp_path / "book.xlsx"))
    sheet = load_workbook(book).active
    assert sheet is not None
    assert sheet["A1"].value == "PV Modules"
    assert sheet["B2"].value == 650


def test_the_compression_level_is_pinned_not_inherited(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Decision 8c pins the level because zlib's default varies across Python
    builds, so an unpinned archive is deterministic on one host and not across
    hosts — and every determinism test here compares two runs on the *same* host,
    which any self-consistent level satisfies.

    Two assertions, because either alone is a tautology. Following the constant
    into the output proves the assignment is live (deleting
    `info._compresslevel = ...` is otherwise invisible: zlib's default happens to
    be 6, the same value). Pinning the constant to 6 proves it is the value
    Decision 8c chose, which following it cannot.
    """
    assert DETERMINISTIC_COMPRESSLEVEL == 6

    def compressed_size(book: Path) -> tuple[int, bytes]:
        with zipfile.ZipFile(book) as archive:
            biggest = max(archive.infolist(), key=lambda i: i.file_size)
            return biggest.compress_size, archive.read(biggest.filename)

    at_default, payload = compressed_size(normalize_archive(_write(tmp_path / "six.xlsx")))
    monkeypatch.setattr(output_module, "DETERMINISTIC_COMPRESSLEVEL", 9)
    at_nine, _ = compressed_size(normalize_archive(_write(tmp_path / "nine.xlsx")))

    def deflated(level: int) -> int:
        engine = zlib.compressobj(level, zlib.DEFLATED, -zlib.MAX_WBITS)
        return len(engine.compress(payload) + engine.flush())

    assert deflated(6) != deflated(9), "fixture too small to distinguish levels"
    assert at_default == deflated(6)
    assert at_nine == deflated(9), "the pinned level is not reaching the ZipInfo"
