"""Executable sanitized-PV path from intake through review and workbook output."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any, LiteralString

import pytest
from openpyxl import load_workbook

from procurement_agent.audit import AuditConnection
from procurement_agent.schema import (
    ConflictClass,
    ConflictStatus,
    Resolution,
    ResolutionAction,
    Severity,
    WorkbookTab,
)
from procurement_agent.services.output.projection import ProjectionPolicy
from procurement_agent.services.vertical_slice import (
    InMemoryClaimStore,
    SanitizedCSVError,
    VerticalSliceResult,
    persist_vertical_slice,
    review_conflict,
    run_sanitized_pv_csv,
)

FIXTURE = Path(__file__).parent / "fixtures" / "ingestion" / "sanitized-pv-module.csv"
INGESTED_AT = datetime(2026, 8, 1, 9, 0, tzinfo=UTC)
DETECTED_AT = datetime(2026, 8, 1, 9, 1, tzinfo=UTC)


def _run(**kwargs: object) -> VerticalSliceResult:
    arguments: dict[str, object] = {
        "ingested_at": INGESTED_AT,
        "detected_at": DETECTED_AT,
        "actor": "integration-test",
    }
    arguments.update(kwargs)
    return run_sanitized_pv_csv(FIXTURE.read_bytes(), **arguments)  # type: ignore[arg-type]


def test_sanitized_csv_reaches_canonical_component_and_conflict_queue() -> None:
    result = _run()

    assert [source.document_id for source in result.sources] == [
        "syn-pv-datasheet",
        "syn-pv-purchase-order",
    ]
    assert len(result.claims) == 3
    assert all(claim.model_config["frozen"] for claim in result.claims)

    component = result.components[0]
    assert component.supplier == "Example Solar Ltd."
    assert component.model == "PV-SAN-650"
    assert component.fields["module_efficiency"][0].value == 22.5
    nameplate = component.fields["nameplate_power"][0]
    assert nameplate.value == 650.0
    assert nameplate.conflict_status is ConflictStatus.OPEN
    assert nameplate.source_ref.extractor_version == "sanitized-pv-csv@1"

    [conflict] = result.conflicts
    assert conflict.conflict_class is ConflictClass.INTER_DOCUMENT
    assert conflict.severity is Severity.HIGH
    assert [candidate.value for candidate in conflict.candidates] == [650.0, 655.0]


def test_claim_store_is_append_only_and_replay_is_idempotent() -> None:
    store = InMemoryClaimStore()
    first = _run(claim_store=store)
    second = _run(claim_store=store)

    assert store.claims == first.claims == second.claims
    assert len(store.claims) == 3


def test_projection_contains_store_policy_and_conflict() -> None:
    projection = _run().workbook_projection(
        policy=ProjectionPolicy(policy_version="test@1", confidence_threshold=0.80)
    )

    assert projection["policy"] == {
        "policy_version": "test@1",
        "confidence_threshold": 0.8,
    }
    conflicts = projection["conflicts"]
    assert isinstance(conflicts, list)
    assert conflicts[0]["field_name"] == "nameplate_power"


def test_persistence_hook_sees_only_a_fully_validated_result() -> None:
    persisted: list[VerticalSliceResult] = []
    result = _run(persist=persisted.append)

    assert persisted == [result]
    assert [event.event_type for event in result.audit_events] == [
        "document_ingested",
        "extraction",
        "document_ingested",
        "extraction",
        "conflict_detected",
    ]
    assert {event.document_id for event in result.audit_events} == {
        "syn-pv-datasheet",
        "syn-pv-purchase-order",
    }


class _Cursor:
    def __init__(self, row: Any = None) -> None:
        self._row = row

    def fetchone(self) -> Any:
        return self._row


class _AuditConnection:
    def __init__(self, *, autocommit: bool = False) -> None:
        self._autocommit = autocommit
        self.calls: list[str] = []
        self.tips: dict[str, tuple[int, bytes]] = {}
        self.commits = 0

    @property
    def autocommit(self) -> bool:
        return self._autocommit

    def execute(self, query: LiteralString, params: object = None) -> _Cursor:
        statement = " ".join(query.split())
        self.calls.append(statement)
        arguments = params if isinstance(params, tuple) else ()
        if statement.startswith("SELECT seq, hash FROM audit.event"):
            return _Cursor(self.tips.get(str(arguments[0])))
        if statement.startswith("INSERT INTO audit.event"):
            self.tips[str(arguments[0])] = (int(arguments[2]), bytes(arguments[4]))
        return _Cursor()

    def commit(self) -> None:
        self.commits += 1


def test_persistence_and_all_audit_intents_share_one_uncommitted_connection() -> None:
    result = _run()
    conn = _AuditConnection()

    def write(connection: AuditConnection, received: VerticalSliceResult) -> str:
        assert received is result
        connection.execute("INSERT VERTICAL SLICE")
        return "stored"

    stored, persisted, events = persist_vertical_slice(conn, result, write=write)

    assert stored == "stored"
    assert persisted.audit_events == ()
    assert len(events) == len(result.audit_events) == 5
    assert [(event.document_id, event.seq) for event in events] == [
        ("syn-pv-datasheet", 0),
        ("syn-pv-datasheet", 1),
        ("syn-pv-purchase-order", 0),
        ("syn-pv-purchase-order", 1),
        ("syn-pv-datasheet", 2),
    ]
    assert conn.calls[0] == "INSERT VERTICAL SLICE"
    assert conn.commits == 0


def test_a_review_after_persistence_emits_only_its_new_resolution_intent() -> None:
    result = _run()
    conn = _AuditConnection()

    _, persisted, _ = persist_vertical_slice(
        conn,
        result,
        write=lambda connection, _: connection.execute("INSERT VERTICAL SLICE"),
    )
    conflict = persisted.conflicts[0]
    resolution = Resolution(
        action=ResolutionAction.SELECT_VALUE,
        resolved_by="procurement.lead",
        resolved_at=datetime(2026, 8, 1, 10, 0, tzinfo=UTC),
        rationale="The signed purchase order supersedes the draft datasheet.",
        value_before=650.0,
        value_after=655.0,
    )
    reviewed = review_conflict(persisted, entry_id=conflict.entry_id, resolution=resolution)

    assert [intent.event_type for intent in reviewed.audit_events] == ["resolution"]
    _, repersisted, events = persist_vertical_slice(
        conn,
        reviewed,
        write=lambda connection, _: connection.execute("INSERT RESOLUTION"),
    )

    assert repersisted.audit_events == ()
    assert [(event.document_id, event.seq, event.event_type) for event in events] == [
        ("syn-pv-purchase-order", 2, "resolution")
    ]


def test_vertical_slice_refuses_autocommit_before_business_write() -> None:
    result = _run()
    conn = _AuditConnection(autocommit=True)
    write_called = False

    def write(_: AuditConnection, __: VerticalSliceResult) -> None:
        nonlocal write_called
        write_called = True

    with pytest.raises(ValueError, match="autocommit=False"):
        persist_vertical_slice(conn, result, write=write)

    assert write_called is False
    assert conn.calls == []


def test_review_selects_existing_candidate_without_rewriting_input() -> None:
    original = _run()
    conflict = original.conflicts[0]
    resolution = Resolution(
        action=ResolutionAction.SELECT_VALUE,
        resolved_by="procurement.lead",
        resolved_at=datetime(2026, 8, 1, 10, 0, tzinfo=UTC),
        rationale="The signed purchase order supersedes the draft datasheet.",
        value_before=650.0,
        value_after=655.0,
    )

    reviewed = review_conflict(original, entry_id=conflict.entry_id, resolution=resolution)

    assert original.conflicts[0].resolution is None
    assert original.components[0].fields["nameplate_power"][0].value == 650.0
    assert reviewed.conflicts[0].resolution == resolution
    field = reviewed.components[0].fields["nameplate_power"][0]
    assert field.value == 655.0
    assert field.conflict_status is ConflictStatus.RESOLVED
    assert field.resolution == resolution
    assert field.source_ref.document_id == "syn-pv-purchase-order"
    assert reviewed.audit_events[-1].event_type == "resolution"

    with pytest.raises(ValueError, match="already has an immutable resolution"):
        review_conflict(reviewed, entry_id=conflict.entry_id, resolution=resolution)


def test_review_refuses_unsourced_override() -> None:
    result = _run()
    resolution = Resolution(
        action=ResolutionAction.ENTER_OVERRIDE,
        resolved_by="procurement.lead",
        resolved_at=datetime(2026, 8, 1, 10, 0, tzinfo=UTC),
        rationale="manual override",
        value_after=660.0,
    )
    with pytest.raises(ValueError, match="selecting an existing value only"):
        review_conflict(result, entry_id=result.conflicts[0].entry_id, resolution=resolution)


def test_writer_emits_all_tabs_provenance_and_open_item(tmp_path: Path) -> None:
    workbook_path = _run().write_workbook(tmp_path / "pv.xlsx", confidence_threshold=0.80)
    workbook = load_workbook(workbook_path)

    assert workbook.sheetnames == [tab.value for tab in WorkbookTab]
    pv = workbook[WorkbookTab.PV_MODULES.value]
    assert pv.max_row == 3
    assert pv["E2"].comment is not None
    assert "syn-pv-datasheet" in pv["E2"].comment.text
    assert pv["H3"].value == "unresolved_conflict"
    open_items = workbook[WorkbookTab.CONFLICTS_OPEN_ITEMS.value]
    assert open_items.max_row == 2
    assert open_items["C2"].value == "nameplate_power"
    assert open_items["D2"].value == "open"


def test_workbook_bytes_are_deterministic(tmp_path: Path) -> None:
    result = _run()
    first = result.write_workbook(tmp_path / "first.xlsx", confidence_threshold=0.80)
    second = result.write_workbook(tmp_path / "second.xlsx", confidence_threshold=0.80)
    assert first.read_bytes() == second.read_bytes()


def test_adapter_rejects_an_off_contract_field_before_persistence() -> None:
    data = FIXTURE.read_text().replace("module_efficiency", "made_up_efficiency").encode()
    called = False

    def persist(_: VerticalSliceResult) -> None:
        nonlocal called
        called = True

    with pytest.raises(SanitizedCSVError, match="not a PV-module contract key"):
        run_sanitized_pv_csv(
            data,
            ingested_at=INGESTED_AT,
            detected_at=DETECTED_AT,
            actor="integration-test",
            persist=persist,
        )
    assert called is False


def test_document_hash_is_independent_of_row_arrival_order() -> None:
    rows = FIXTURE.read_bytes().splitlines()
    header, *body = rows
    reordered = b"\n".join([header, *reversed(body), b""])

    original = _run()
    permuted = run_sanitized_pv_csv(
        reordered,
        ingested_at=INGESTED_AT,
        detected_at=DETECTED_AT,
        actor="integration-test",
    )

    assert {source.document_id: source.content_hash for source in original.sources} == {
        source.document_id: source.content_hash for source in permuted.sources
    }
