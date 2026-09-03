"""Comparison / Output (FR-OUT-01 .. FR-OUT-06).

Canonical schema -> multi-tab Excel writer with provenance and conditional
formatting.
"""

import json
import os
import re
import tempfile
import zipfile
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ...schema import (
    CATEGORY_TO_TAB,
    CanonicalField,
    CellFlag,
    ComponentInstance,
    ConflictStatus,
    WorkbookTab,
    encode_value,
)

#: plan.md Decision 8c. 1980-01-01 is the ZIP format's epoch floor; noon keeps
#: every member clear of it under any timezone interpretation.
DETERMINISTIC_TIMESTAMP = (1980, 1, 1, 12, 0, 0)

#: Replaces openpyxl's `Microsoft Excel Compatible / Openpyxl <version>`, which
#: embeds the library version in `docProps/app.xml` and so changes the output
#: bytes on a patch bump with zero data change.
DETERMINISTIC_APPLICATION = "Procurement Agent"

#: plan.md Decision 8c: sorted entries with `[Content_Types].xml` forced first.
#: openpyxl 3.1.5 writes it *last*, so preserving `infolist()` order preserves a
#: non-conformant archive.
CONTENT_TYPES = "[Content_Types].xml"

#: Decision 8c. Pinned so output bytes do not depend on the host's zlib default,
#: which varies across Python builds.
DETERMINISTIC_COMPRESSLEVEL = 6

#: Decision 8c: always "Unix", else `ZipInfo.__init__` picks 0 on Windows and 3
#: elsewhere, so the same store would produce different bytes per platform.
_CREATE_SYSTEM_UNIX = 3
_EXTERNAL_ATTR = 0o644 << 16

_APPLICATION_RE = re.compile(rb"<Application>[^<]*</Application>")
_APP_VERSION_RE = re.compile(rb"<AppVersion>[^<]*</AppVersion>")
_DCTERMS_RE = re.compile(rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:)")

#: The epoch as it appears inside `docProps/core.xml`, matching the ZIP stamp.
#: Substituted with \g<N> back-references, not \N - the replacement starts with
#: "1980", so "\1" + "1980..." would parse as group 11.
_CORE_XML_EPOCH = b"1980-01-01T12:00:00Z"


def normalize_archive(path: Path) -> Path:
    """Strip every wall-clock, library-version and platform trace from an `.xlsx`.

    Required by FR-OUT-06 and AC-7. Freezing `workbook.properties.modified` is
    **not sufficient**: openpyxl re-stamps it unconditionally on save
    (`openpyxl/writer/excel.py:292`, no opt-out), and each ZIP local header
    carries an mtime derived from the clock. See issue #13.

    Five sources of run-to-run variance, all of which have to go - missing any
    one leaves the archive non-deterministic while looking fixed:

    1. ZIP local-header mtimes, from the clock via `writestr`.
    2. `docProps/core.xml` `dcterms:created` / `dcterms:modified`. This is the one
       an earlier version of this function missed, so two saves seconds apart
       still differed while every other member was byte-identical.
    3. `docProps/app.xml`, which embeds the openpyxl version verbatim.
    4. Compression level. Decision 8c warns by name that `ZipFile(compresslevel=)`
       is **silently ignored** for a hand-built `ZipInfo`. `writestr` takes its
       own `compresslevel` and is the public setter.
    5. `create_system` and `external_attr`, which otherwise vary by platform.

    Member order is normalised too: sorted, with `[Content_Types].xml` first per
    Decision 8c. openpyxl writes it last.
    """
    with zipfile.ZipFile(path) as source:
        members = {info.filename: source.read(info.filename) for info in source.infolist()}

    ordered = sorted(members, key=lambda name: (name != CONTENT_TYPES, name))

    # `mkstemp`, not pid: two callers in one process share a pid, and the loser
    # of the race raised FileNotFoundError out of a call that had succeeded.
    handle, staging_name = tempfile.mkstemp(
        dir=path.parent, prefix=f"{path.name}.", suffix=".normalizing"
    )
    os.close(handle)
    staging = Path(staging_name)
    try:
        with zipfile.ZipFile(staging, "w", zipfile.ZIP_DEFLATED) as target:
            for name in ordered:
                payload = members[name]
                # OPC part names are case-insensitive; openpyxl always writes
                # this casing, but a foreign workbook need not.
                lowered = name.casefold()
                if lowered == "docprops/app.xml":
                    payload = _APPLICATION_RE.sub(
                        f"<Application>{DETERMINISTIC_APPLICATION}</Application>".encode(), payload
                    )
                    payload = _APP_VERSION_RE.sub(b"<AppVersion>1.0</AppVersion>", payload)
                elif lowered == "docprops/core.xml":
                    payload = _DCTERMS_RE.sub(rb"\g<1>" + _CORE_XML_EPOCH + rb"\g<2>", payload)

                info = zipfile.ZipInfo(name, date_time=DETERMINISTIC_TIMESTAMP)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = _CREATE_SYSTEM_UNIX
                info.external_attr = _EXTERNAL_ATTR
                # Decision 8c warns that `ZipFile(compresslevel=)` is ignored for
                # a hand-built ZipInfo. `writestr(..., compresslevel=)` is the
                # public setter; it writes the private attribute itself, including
                # the 3.13 rename `_compress_level`.
                target.writestr(info, payload, compresslevel=DETERMINISTIC_COMPRESSLEVEL)
        # os.replace, not shutil.move: same directory, so this is an atomic
        # rename. shutil.move falls back to copy-then-unlink across filesystems,
        # and a copy that dies partway leaves the destination truncated - at
        # which point unlinking the staging file destroys the only intact copy.
        os.replace(staging, path)
    except BaseException:
        staging.unlink(missing_ok=True)
        raise
    return path


def flags_for(field: CanonicalField, *, confidence_threshold: float) -> set[CellFlag]:
    """Which of the four conditional-formatting states apply to a cell (FR-OUT-04).

    A cell can carry more than one: a web-supplemented value can also be
    low-confidence.
    """
    flags: set[CellFlag] = set()
    if field.is_missing():
        flags.add(CellFlag.MISSING_DATA)
    if field.is_web_supplemented():
        flags.add(CellFlag.WEB_SUPPLEMENTED)
    if field.confidence < confidence_threshold:
        flags.add(CellFlag.LOW_CONFIDENCE)
    if field.conflict_status in {ConflictStatus.OPEN, ConflictStatus.INSUFFICIENT_EVIDENCE}:
        flags.add(CellFlag.UNRESOLVED_CONFLICT)
    return flags


def write_workbook(
    components: list[ComponentInstance],
    destination: Path,
    *,
    suppliers_as_rows: bool = True,
    confidence_threshold: float,
) -> Path:
    """Emit the workbook. All thirteen tabs, always (FR-OUT-02, AC-3).

    Must be deterministically regenerable from the canonical store (FR-OUT-06),
    which means no timestamps or ordering derived from anything but the store
    itself, plus an explicit generated-on stamp and per-source data vintage.

    Every comparison cell carries provenance via cell comment, a link into the
    Sources tab, or an adjacent column (FR-OUT-03). No unsourced values
    (NFR-01, AC-4).
    """
    if not suppliers_as_rows:
        raise NotImplementedError(
            "the initial deterministic writer implements suppliers_as_rows=True only"
        )
    if not 0.0 <= confidence_threshold <= 1.0:
        raise ValueError("confidence_threshold must be between 0 and 1")

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    initial = workbook.active
    assert initial is not None
    workbook.remove(initial)

    sheets = {tab: workbook.create_sheet(tab.value) for tab in expected_tabs()}
    category_headers = [
        "Supplier",
        "Model",
        "Field",
        "Condition",
        "Value",
        "Unit",
        "Confidence",
        "Flags",
        "Provenance",
    ]
    for tab in CATEGORY_TO_TAB.values():
        sheet = sheets[tab]
        sheet.append(category_headers)
        _style_header(sheet[1])
        sheet.freeze_panes = "A2"

    ordered = sorted(components, key=ComponentInstance.ordering_key)
    for component in ordered:
        sheet = sheets[CATEGORY_TO_TAB[component.component_category]]
        for field_name in sorted(component.fields):
            for field in sorted(component.fields[field_name], key=_field_sort_key):
                cell_flags = flags_for(field, confidence_threshold=confidence_threshold)
                provenance = _provenance_text(field)
                row = [
                    _safe_excel_text(component.supplier),
                    _safe_excel_text(component.model),
                    field_name,
                    _display(field.condition),
                    _display(field.value),
                    field.unit,
                    field.confidence,
                    ", ".join(sorted(flag.value for flag in cell_flags)),
                    provenance,
                ]
                sheet.append(row)
                value_cell = sheet.cell(row=sheet.max_row, column=5)
                value_cell.comment = Comment(provenance, DETERMINISTIC_APPLICATION)
                _apply_flag_fill(value_cell, cell_flags)

    _write_summary(sheets[WorkbookTab.EXECUTIVE_SUMMARY], ordered)
    _write_open_items(sheets[WorkbookTab.CONFLICTS_OPEN_ITEMS], ordered)
    _write_provenance(sheets[WorkbookTab.SOURCES_PROVENANCE], ordered)
    _write_cross_cutting(sheets[WorkbookTab.COMPLIANCE_MATRIX], ordered, "compliance")
    _write_cross_cutting(sheets[WorkbookTab.TAX_INCENTIVES], ordered, "tax")

    for sheet in workbook.worksheets:
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(sheet.columns, start=1):
            letter = get_column_letter(index)
            width = min(60, max(12, *(len(str(cell.value or "")) + 2 for cell in column)))
            sheet.column_dimensions[letter].width = width

    workbook.save(destination)
    return normalize_archive(destination)


_FLAG_FILLS: dict[CellFlag, PatternFill] = {
    CellFlag.UNRESOLVED_CONFLICT: PatternFill("solid", fgColor="F4CCCC"),
    CellFlag.MISSING_DATA: PatternFill("solid", fgColor="D9D9D9"),
    CellFlag.LOW_CONFIDENCE: PatternFill("solid", fgColor="FFF2CC"),
    CellFlag.WEB_SUPPLEMENTED: PatternFill("solid", fgColor="CFE2F3"),
}


def _style_header(cells: Iterable[Cell | MergedCell]) -> None:
    """Apply one deterministic, restrained header style."""
    for cell in cells:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def _display(value: object) -> str:
    return json.dumps(
        encode_value(value), sort_keys=True, ensure_ascii=False, separators=(",", ":")
    )


def _safe_excel_text(value: str) -> str:
    """Keep untrusted text from becoming a spreadsheet formula.

    Supplier and model names originate outside the application.  Excel and
    LibreOffice interpret cells beginning with ``=``, ``+``, ``-`` or ``@`` as
    formulas, and tab/newline prefixes are also used to bypass shallow checks.
    A leading apostrophe forces text semantics while preserving the visible
    content in spreadsheet applications.
    """
    if value.startswith(("=", "+", "-", "@", "\t", "\r", "\n")):
        return "'" + value
    return value


def _field_sort_key(field: CanonicalField) -> tuple[str, str, str]:
    return (_display(field.condition), _display(field.value), _display(field.source_ref))


def _provenance_text(field: CanonicalField) -> str:
    ref = field.source_ref
    if ref.document_id is not None:
        location = ref.document_id
        if ref.page is not None:
            location += f" page {ref.page}"
        if ref.section is not None:
            location += f" [{ref.section}]"
    else:
        location = ref.url or ""
    extractor = f"; extractor={ref.extractor_version}" if ref.extractor_version else ""
    return f"{field.source_tier.value}: {location}{extractor}"


def _apply_flag_fill(cell: Cell, flags: set[CellFlag]) -> None:
    # More severe review states win visually; every state remains present in the Flags column.
    precedence = (
        CellFlag.UNRESOLVED_CONFLICT,
        CellFlag.MISSING_DATA,
        CellFlag.LOW_CONFIDENCE,
        CellFlag.WEB_SUPPLEMENTED,
    )
    for flag in precedence:
        if flag in flags:
            cell.fill = _FLAG_FILLS[flag]
            return


def _write_summary(sheet: Worksheet, components: list[ComponentInstance]) -> None:
    sheet.append(["Metric", "Value"])
    _style_header(sheet[1])
    sheet.append(["Component instances", len(components)])
    sheet.append(["Suppliers", len({component.supplier for component in components})])
    sheet.append(
        [
            "Unresolved fields",
            sum(len(component.unresolved_conflicts()) for component in components),
        ]
    )


def _write_open_items(sheet: Worksheet, components: list[ComponentInstance]) -> None:
    sheet.append(["Supplier", "Model", "Field", "Status", "Provenance"])
    _style_header(sheet[1])
    for component in components:
        for field_name in component.unresolved_conflicts():
            for field in component.fields[field_name]:
                if CellFlag.UNRESOLVED_CONFLICT in flags_for(field, confidence_threshold=0.0):
                    sheet.append(
                        [
                            _safe_excel_text(component.supplier),
                            _safe_excel_text(component.model),
                            field_name,
                            field.conflict_status.value,
                            _provenance_text(field),
                        ]
                    )


def _write_provenance(sheet: Worksheet, components: list[ComponentInstance]) -> None:
    sheet.append(["Supplier", "Model", "Field", "Source tier", "Source"])
    _style_header(sheet[1])
    rows = {
        (
            _safe_excel_text(component.supplier),
            _safe_excel_text(component.model),
            field_name,
            field.source_tier.value,
            _provenance_text(field),
        )
        for component in components
        for field_name, fields in component.fields.items()
        for field in fields
    }
    for row in sorted(rows):
        sheet.append(row)


def _write_cross_cutting(sheet: Worksheet, components: list[ComponentInstance], topic: str) -> None:
    sheet.append(["Supplier", "Model", "Field", "Value", "Provenance"])
    _style_header(sheet[1])
    tax_keys = {
        "domestic_content_status",
        "domestic_content_percentage",
        "baba_status",
        "feoc_pfe_status",
    }
    for component in components:
        for field_name in sorted(component.fields):
            include = field_name in tax_keys if topic == "tax" else field_name not in tax_keys
            if not include:
                continue
            for field in component.fields[field_name]:
                sheet.append(
                    [
                        _safe_excel_text(component.supplier),
                        _safe_excel_text(component.model),
                        field_name,
                        _display(field.value),
                        _provenance_text(field),
                    ]
                )


def expected_tabs() -> list[WorkbookTab]:
    """The thirteen tabs in order, for a writer to iterate (FR-OUT-02).

    Pinned by `test_expected_tabs_returns_all_thirteen_in_order`. AC-3 itself is
    *not* asserted here and cannot be until `write_workbook` exists - AC-3 wants
    thirteen tabs in a generated workbook with conditional formatting, and this
    is only the order they go in.
    """
    return list(WorkbookTab)
